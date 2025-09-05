import re
from typing import IO
import zoneinfo
from openpyxl import load_workbook
import datetime

import openpyxl.utils
from ..schedule import Schedule, ScheduleEntry
from ..schedule_utils import LAB
from ..jwapi_model import ErrorEntry
from .api_model import PhxpLabCourse, PhxpResponse


def arrange(
    in_file: str | IO[bytes],
    out_file: str,
    schedule: Schedule,
    error_entries: list[ErrorEntry] | None = None,
):
    # Load workbook and worksheet
    wb = load_workbook(in_file)
    ws = wb.active

    if ws is None:
        raise ValueError("出错了：无法加载活动工作表")

    # Insert new conflict course column
    original_max_col = ws.max_column
    new_col_idx = original_max_col + 1
    ws.insert_cols(new_col_idx)

    # Initialize variables to track merged cell values
    last_week_id = None
    last_date = None
    last_day_of_week = None

    header_row = 1
    while (
        header_row <= ws.max_row
        and type(ws.cell(header_row + 1, 1).value) != int
        and "周次" not in str(ws.cell(header_row, 1).value)
    ):
        header_row += 1

    # Process each row
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Handle merged cells by carrying forward values from above
        # Week ID (column 1)
        current_week_id = ws.cell(row=row_idx, column=1).value
        if current_week_id is not None:
            last_week_id = current_week_id
        else:
            current_week_id = last_week_id

        # Date (column 2)
        current_date = ws.cell(row=row_idx, column=2).value
        if current_date is not None:
            last_date = current_date
        else:
            current_date = last_date

        # Day of week (column 3)
        current_day_of_week = ws.cell(row=row_idx, column=3).value
        if current_day_of_week is not None:
            last_day_of_week = current_day_of_week
        else:
            current_day_of_week = last_day_of_week

        try:
            # Validate required fields
            if None in (current_week_id, current_date, current_day_of_week):
                raise ValueError("Missing index values in merged cells")

            # Convert day of week to number
            day_mapping = {
                "周一": 1,
                "周二": 2,
                "周三": 3,
                "周四": 4,
                "周五": 5,
                "周六": 6,
                "周日": 7,
            }
            q_day_of_week = day_mapping.get(str(current_day_of_week))
            if q_day_of_week is None:
                raise ValueError(f"无效的星期格式: {current_day_of_week}")

            # Parse date
            if isinstance(current_date, str):
                q_date = datetime.datetime.strptime(current_date, "%Y.%m.%d")
            elif isinstance(current_date, datetime.datetime):
                q_date = current_date  # Assume datetime object
            else:
                raise ValueError(f"无效的日期格式: {current_date}")

            # Parse time span
            time_span = ws.cell(row=row_idx, column=4).value
            time_match = re.match(r"(\d+)(-|、)(\d+)", str(time_span))
            if not time_match:
                raise ValueError(f"无效的时间段格式: {time_span}")
            q_time_span = (int(time_match.group(1)), int(time_match.group(3)))

            # Query for conflicts
            conflicts = schedule.query_lesson_at(
                int(str(current_week_id)), q_date, q_day_of_week, q_time_span
            )
            conflict_names = [e.lab_name or e.name for e in conflicts]
            _ = ws.cell(row=row_idx, column=new_col_idx, value="，".join(conflict_names))

        except ValueError as e:
            if error_entries is not None:
                error_entries.append(
                    ErrorEntry(
                        entry=f"第 {row_idx} 行："
                        + "\t".join(
                            str(ws.cell(row=row_idx, column=col).value)
                            for col in range(1, ws.max_column + 1)
                        ),
                        reason=str(e),
                    )
                )
            continue

    _ = ws.cell(row=header_row, column=new_col_idx, value="冲突课程")
    col_letter = openpyxl.utils.get_column_letter(new_col_idx)
    ws.auto_filter.ref = f"{col_letter}{header_row}:{col_letter}{ws.max_row}"

    wb.save(out_file)


def parse_lab_entry(item: PhxpLabCourse):
    zone = zoneinfo.ZoneInfo("Asia/Shanghai")

    def _add_tz(t: str):
        return datetime.datetime.strptime(t, "%H:%M").time().replace(tzinfo=zone)

    return ScheduleEntry(
        item.ModuleName,
        datetime.datetime.strptime(item.ClassDate, "%Y/%m/%d %H:%M:%S").date(),
        [(_add_tz(item.StartTime), _add_tz(item.EndTime))],
        item.ClassRoom,
        LAB,
        teacher=item.TeacherName,
        lab_name=item.LabName,
    )


def create_schedule_from(
    obj: PhxpResponse, semester_desc: str, start_date: datetime.date
):
    entries: list[ScheduleEntry] = []
    for item in obj.rows:
        entry = parse_lab_entry(item)
        # if entry is None:
        #     raise ValueError(f"遇到无法解析的课表条目。")
        entries.append(entry)
    return Schedule(entries, semester_desc, start_date)
